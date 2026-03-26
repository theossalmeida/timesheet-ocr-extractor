from __future__ import annotations
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models.timesheet import ExtractionResult, TimesheetRow

OCORRENCIA_COLORS: dict[str, str] = {
    "ferias": "FFF9C4",
    "feriado": "C8E6C9",
    "falta_justificada": "FFE0B2",
    "falta_injustificada": "FFCDD2",
    "licenca_medica": "E1BEE7",
    "afastamento": "B3E5FC",
    "folga": "F0F4C3",
    "dsr": "D7CCC8",
    "meio_periodo": "FCE4EC",
    "trabalho_normal": "FFFFFF",
    "outro": "F5F5F5",
}

TIPO_LABELS: dict[str, str] = {
    "ferias": "Férias",
    "feriado": "Feriado",
    "falta_justificada": "Falta Justificada",
    "falta_injustificada": "Falta Injustificada",
    "licenca_medica": "Licença Médica",
    "afastamento": "Afastamento",
    "folga": "Folga",
    "dsr": "DSR",
    "meio_periodo": "Meio Período",
    "trabalho_normal": "Trabalho Normal",
    "outro": "Outro",
}

_THIN = Side(border_style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

HEADERS = ["Data", "Entrada 1", "Saída 1", "Entrada 2", "Saída 2", "Ocorrência", "Tipo"]
COL_WIDTHS = [12, 10, 10, 10, 10, 22, 20]


def _apply_header(ws) -> None:
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
        ws.column_dimensions[cell.column_letter].width = width


def _row_fill(row: TimesheetRow, row_idx: int) -> PatternFill:
    if row.ocorrencia_tipo and row.ocorrencia_tipo in OCORRENCIA_COLORS:
        color = OCORRENCIA_COLORS[row.ocorrencia_tipo]
        if color != "FFFFFF":
            return PatternFill("solid", fgColor=color)
    return _ALT_FILL if row_idx % 2 == 0 else _WHITE_FILL


def build_excel(result: ExtractionResult) -> bytes:
    wb = openpyxl.Workbook()

    # --- Tab 1: Registros de Ponto ---
    ws1 = wb.active
    ws1.title = "Registros de Ponto"
    _apply_header(ws1)

    for i, row in enumerate(result.rows, start=1):
        excel_row = i + 1
        fill = _row_fill(row, i)
        values = [
            row.data,
            row.entrada_1,
            row.saida_1,
            row.entrada_2,
            row.saida_2,
            row.ocorrencia_raw,
            TIPO_LABELS.get(row.ocorrencia_tipo or "", "") if row.ocorrencia_tipo else "",
        ]
        alignments = [_LEFT, _CENTER, _CENTER, _CENTER, _CENTER, _LEFT, _LEFT]
        for col, (val, align) in enumerate(zip(values, alignments), start=1):
            cell = ws1.cell(row=excel_row, column=col, value=val)
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = align

    # --- Tab 2: Resumo ---
    ws2 = wb.create_sheet("Resumo")
    dates = [r.data for r in result.rows if r.data]
    trabalhados = sum(1 for r in result.rows if r.entrada_1)

    tipo_counts: dict[str, int] = {}
    for r in result.rows:
        key = r.ocorrencia_tipo or "trabalho_normal"
        tipo_counts[key] = tipo_counts.get(key, 0) + 1

    summary_data = [
        ("Total de registros", len(result.rows)),
        ("Data inicial", min(dates) if dates else "—"),
        ("Data final", max(dates) if dates else "—"),
        ("Dias trabalhados", trabalhados),
        ("Provider usado", result.provider),
        ("Tipo de PDF", result.pdf_type),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Ocorrências", ""),
    ]
    for tipo, count in sorted(tipo_counts.items()):
        summary_data.append((TIPO_LABELS.get(tipo, tipo), count))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        if label in ("Ocorrências", "") and value == "":
            if label == "Ocorrências":
                cell_a.font = header_font
                cell_a.fill = header_fill
                cell_b.fill = header_fill
        cell_a.border = _BORDER
        cell_b.border = _BORDER

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 28

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
