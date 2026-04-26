from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MONTHS_PT = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
COLUMNS_LETTER = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']

FORMULAS = {
    'Total Apurado': '=SUM({cell1}:{cell2})',
    'Descontos': '',
    'Base FGTS': '={cell1}-{cell2}',
    'FGTS Devido': '={cell1}*8%',
    'FGTS Pago': '',
    'Diferença FGTS': '={cell1}-{cell2}',
}

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=13, color="1E3A5F")
_DESC_FONT = Font(size=10)
_VALUE_FONT = Font(size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
_NO_FILL = PatternFill(fill_type=None)

_THIN = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Number of empty rows between year sections
_GAP_ROWS = 5


def build_contracheque_excel(
    salary_data: dict[str, dict[str, dict[str, float]]]
) -> bytes:
    """
    Build an Excel workbook from aggregated salary data.

    Args:
        salary_data: Nested dict structured as
            {year_str: {month_num_str: {descricao: valor}}}
            e.g. {"2022": {"1": {"Salário Básico": 10568.88, "Anuênio": 200.0}}}

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Salarial"

    current_row = 1

    for year_idx, year in enumerate(sorted(salary_data.keys())):
        months_data = salary_data[year]

        # Collect descriptions in insertion order (first seen month first)
        descriptions: list[str] = []
        seen: set[str] = set()
        for month_key in sorted(months_data.keys(), key=int):
            for desc in months_data[month_key]:
                if desc not in seen:
                    seen.add(desc)
                    descriptions.append(desc)

        if not descriptions:
            continue

        formulas_descriptions = ["", "Total Apurado", "Descontos", "Base FGTS", "FGTS Devido", "FGTS Pago", "Diferença FGTS"]
        for formula_desc in formulas_descriptions: descriptions.append(formula_desc) 

        # ── Title row ────────────────────────────────────────────────
        title_cell = ws.cell(row=current_row, column=1, value=f"Ficha Salarial {year}")
        title_cell.font = _TITLE_FONT
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 22
        # Merge across description col + 12 month cols
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=13,
        )
        current_row += 1

        # ── Header row ───────────────────────────────────────────────
        desc_header = ws.cell(row=current_row, column=1, value="Descrição")
        desc_header.font = _HEADER_FONT
        desc_header.fill = _HEADER_FILL
        desc_header.alignment = Alignment(horizontal="left", vertical="center")
        desc_header.border = _THIN_BORDER

        for month_idx, month_name in enumerate(MONTHS_PT):
            col = month_idx + 2
            cell = ws.cell(row=current_row, column=col, value=month_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER

        ws.row_dimensions[current_row].height = 18
        current_row += 1
        initial_row = current_row
        # ── Data rows ────────────────────────────────────────────────
        for row_idx, desc in enumerate(descriptions):
            fill = _ALT_FILL if row_idx % 2 == 0 else _NO_FILL

            desc_cell = ws.cell(row=current_row, column=1, value=desc)
            desc_cell.font = _DESC_FONT
            desc_cell.fill = fill
            desc_cell.alignment = Alignment(horizontal="left", vertical="center")
            desc_cell.border = _THIN_BORDER

            for month_idx in range(12):
                month_key = str(month_idx + 1)

                if desc == 'Total Apurado':
                    column_letter = COLUMNS_LETTER[month_idx+1]
                    valor = f'=SUM({column_letter}{initial_row}:{column_letter}{current_row-1})'

                elif desc == 'Diferença FGTS' or desc == 'Base FGTS':
                    column_letter = COLUMNS_LETTER[month_idx+1]
                    valor = f'={column_letter}{current_row-2}-{column_letter}{current_row-1}'

                elif desc == 'FGTS Devido':
                    column_letter = COLUMNS_LETTER[month_idx+1]
                    valor = f'={column_letter}{current_row-1}*8%'

                else: 
                    valor = months_data.get(month_key, {}).get(desc)
                col = month_idx + 2
                cell = ws.cell(row=current_row, column=col, value=valor)
                cell.font = _VALUE_FONT
                cell.fill = fill
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = _THIN_BORDER
                if valor is not None:
                    cell.number_format = "#,##0.00"

            current_row += 1

        # ── Gap rows between year sections ───────────────────────────
        current_row += _GAP_ROWS

    # ── Column widths ────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, 14):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
