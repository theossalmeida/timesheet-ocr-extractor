from __future__ import annotations

import io
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.frequency_cycle_service import ClassifiedDay

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_VALUE_FONT = Font(size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
_NO_FILL = PatternFill(fill_type=None)
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_frequency_cycle_excel(rows: list[ClassifiedDay], provider: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ciclos"

    headers = ["Data", "Dia", "Situacao", "Escala", "Marcadores PDF", "Pagina PDF"]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    for row_idx, item in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _NO_FILL
        values = [
            item.date.strftime("%d/%m/%Y"),
            item.cycle_day,
            item.situation,
            item.scale,
            item.details,
            item.page,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _VALUE_FONT
            cell.fill = fill
            cell.border = _BORDER
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (3, 5) else "center",
                vertical="center",
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    widths = [14, 8, 30, 10, 24, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    summary = wb.create_sheet("Resumo")
    summary.append(["Campo", "Valor"])
    summary.append(["Provider", provider])
    summary.append(["Dias classificados", len(rows)])
    if rows:
        summary.append(["Data inicial", rows[0].date.strftime("%d/%m/%Y")])
        summary.append(["Data final", rows[-1].date.strftime("%d/%m/%Y")])
    summary.append([])
    summary.append(["Situacao", "Quantidade"])
    for situation, count in Counter(row.situation for row in rows).items():
        summary.append([situation, count])

    for col_idx in range(1, 3):
        summary.column_dimensions[get_column_letter(col_idx)].width = 28
    for row in summary.iter_rows():
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
