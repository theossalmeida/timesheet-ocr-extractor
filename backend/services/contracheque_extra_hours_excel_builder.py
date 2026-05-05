from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_VALUE_FONT = Font(size=10)
_ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
_NO_FILL = PatternFill(fill_type=None)

_THIN = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _month_sort_key(competencia: str) -> tuple[int, int]:
    try:
        month, year = competencia.split("/")
        return int(year), int(month)
    except (ValueError, AttributeError):
        return 9999, 99


def build_contracheque_extra_hours_excel(
    extra_hours_data: dict[str, dict[str, float]],
    columns: list[str],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Horas Extras"

    headers = ["Data", *columns, "Total"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_idx, competencia in enumerate(sorted(extra_hours_data, key=_month_sort_key), start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else _NO_FILL

        date_cell = ws.cell(row=row_idx, column=1, value=competencia)
        date_cell.font = _VALUE_FONT
        date_cell.fill = fill
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = _THIN_BORDER

        for col_idx, description in enumerate(columns, start=2):
            value = extra_hours_data[competencia].get(description)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _VALUE_FONT
            cell.fill = fill
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _THIN_BORDER
            if value is not None:
                cell.number_format = "#,##0.00"

        first_value_col = get_column_letter(2)
        last_value_col = get_column_letter(len(columns) + 1)
        total_cell = ws.cell(
            row=row_idx,
            column=len(headers),
            value=f"=SUM({first_value_col}{row_idx}:{last_value_col}{row_idx})",
        )
        total_cell.font = Font(bold=True, size=10)
        total_cell.fill = fill
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = _THIN_BORDER
        total_cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 12
    for col_idx, header in enumerate(headers[1:], start=2):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 14), 32)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
