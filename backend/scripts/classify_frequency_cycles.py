from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from services.frequency_cycle_service import (  # noqa: E402
    ClassifiedDay,
    classify_frequency_days,
    compare_with_expected,
    extract_frequency_days_pdfplumber,
    parse_excel_date,
)


def load_expected_excel(excel_path: Path, sheet_name: str = "Ciclos") -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    expected = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_date = parse_excel_date(row[0])
        if row_date is None or not row[2]:
            continue
        cycle_day = int(row[1]) if isinstance(row[1], (int, float)) else None
        expected[row_date] = (cycle_day, str(row[2]).strip())

    return expected


def write_xlsx(rows: list[ClassifiedDay], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Classificacao"
    headers = [
        "Data",
        "Dia",
        "Situacao calculada",
        "Situacao base",
        "Escala",
        "Marcadores PDF",
        "Pagina PDF",
        "Dia esperado",
        "Situacao esperada",
        "Match exato",
        "Match base",
        "Linha PDF",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.date.strftime("%d/%m/%Y"),
            row.cycle_day,
            row.situation,
            row.core_situation,
            row.scale,
            row.details,
            row.page,
            row.expected_cycle_day,
            row.expected_situation,
            row.exact_match,
            row.core_match,
            row.pdf_line,
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    wb.save(output_path)


def write_csv(rows: list[ClassifiedDay], output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle, delimiter=";")
        writer.writerow([
            "Data",
            "Dia",
            "Situacao calculada",
            "Situacao base",
            "Escala",
            "Marcadores PDF",
            "Pagina PDF",
            "Dia esperado",
            "Situacao esperada",
            "Match exato",
            "Match base",
            "Linha PDF",
        ])
        for row in rows:
            writer.writerow([
                row.date.strftime("%d/%m/%Y"),
                row.cycle_day,
                row.situation,
                row.core_situation,
                row.scale,
                row.details,
                row.page,
                row.expected_cycle_day,
                row.expected_situation,
                row.exact_match,
                row.core_match,
                row.pdf_line,
            ])


def print_summary(rows: list[ClassifiedDay]) -> None:
    print(f"Classified days: {len(rows)}")
    if not rows:
        return

    print(f"Date range: {rows[0].date:%d/%m/%Y} to {rows[-1].date:%d/%m/%Y}")
    comparable = [row for row in rows if row.exact_match is not None]
    if comparable:
        exact_matches = sum(1 for row in comparable if row.exact_match)
        core_matches = sum(1 for row in comparable if row.core_match)
        print(f"Compared with Excel: {len(comparable)} days")
        print(f"Exact matches: {exact_matches}/{len(comparable)}")
        print(f"Core matches: {core_matches}/{len(comparable)}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Classify Petrobras frequency-report days into embarked/off-cycle situations."
    )
    parser.add_argument("pdf", type=Path, help="Frequency PDF to classify")
    parser.add_argument("--expected-excel", type=Path, help="Optional Excel file to compare against")
    parser.add_argument("--sheet", default="Ciclos", help="Expected Excel sheet name")
    parser.add_argument("--output-xlsx", type=Path, help="Optional XLSX output path")
    parser.add_argument("--output-csv", type=Path, help="Optional CSV output path")
    args = parser.parse_args()

    frequency_days = extract_frequency_days_pdfplumber(args.pdf.read_bytes())
    classified = classify_frequency_days(frequency_days)

    if args.expected_excel:
        classified = compare_with_expected(
            classified,
            load_expected_excel(args.expected_excel, args.sheet),
        )

    if args.output_xlsx:
        write_xlsx(classified, args.output_xlsx)
    if args.output_csv:
        write_csv(classified, args.output_csv)

    print_summary(classified)


if __name__ == "__main__":
    main()
