import io
from datetime import date

import openpyxl

from services.frequency_cycle_excel_builder import build_frequency_cycle_excel
from services.frequency_cycle_service import ClassifiedDay


def _classified(day: int, ocr_corrected: bool) -> ClassifiedDay:
    return ClassifiedDay(
        date=date(2022, 1, day),
        cycle_day=1,
        situation="EMBARCADO",
        core_situation="EMBARCADO",
        scale="HS02",
        details="",
        pdf_line="",
        page=1,
        ocr_corrected=ocr_corrected,
    )


def test_ocr_corrected_rows_are_flagged_in_new_column():
    rows = [_classified(1, ocr_corrected=False), _classified(2, ocr_corrected=True)]
    wb = openpyxl.load_workbook(io.BytesIO(build_frequency_cycle_excel(rows, "tesseract")))
    ws = wb["Ciclos"]

    headers = [cell.value for cell in ws[1]]
    col = headers.index("Preenchida por Aproximacao") + 1

    assert ws.cell(row=2, column=col).value is None
    assert ws.cell(row=3, column=col).value == "Sim"


def test_resumo_counts_approximated_days():
    rows = [_classified(1, ocr_corrected=False), _classified(2, ocr_corrected=True)]
    wb = openpyxl.load_workbook(io.BytesIO(build_frequency_cycle_excel(rows, "tesseract")))
    summary = wb["Resumo"]

    values = {row[0].value: row[1].value for row in summary.iter_rows(min_row=1, max_row=10) if row[0].value}
    assert values["Dias preenchidos por aproximacao"] == 1
