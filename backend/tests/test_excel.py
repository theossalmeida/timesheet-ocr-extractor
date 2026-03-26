import io
import openpyxl
import pytest
from models.timesheet import ExtractionResult, TimesheetRow
from services.excel_builder import build_excel, OCORRENCIA_COLORS


def _make_result(rows=None) -> ExtractionResult:
    if rows is None:
        rows = [
            TimesheetRow(data="01/03/2024", entrada_1="08:00", saida_1="17:00"),
            TimesheetRow(data="04/03/2024", ocorrencia_raw="FERIAS", ocorrencia_tipo="ferias"),
            TimesheetRow(data="05/03/2024", entrada_1="08:00", saida_1="12:00", ocorrencia_tipo="meio_periodo"),
        ]
    return ExtractionResult(rows=rows, provider="gemini", pdf_type="native")


def test_returns_bytes():
    result = _make_result()
    output = build_excel(result)
    assert isinstance(output, bytes)
    assert len(output) > 0


def test_two_tabs():
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(_make_result())))
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "Registros de Ponto"
    assert wb.sheetnames[1] == "Resumo"


def test_header_row_bold():
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(_make_result())))
    ws = wb["Registros de Ponto"]
    assert ws.cell(row=1, column=1).font.bold is True


def test_header_background_color():
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(_make_result())))
    ws = wb["Registros de Ponto"]
    fill = ws.cell(row=1, column=1).fill
    assert fill.fgColor.rgb.upper().endswith("1E3A5F")


def test_ferias_row_color():
    rows = [TimesheetRow(data="04/03/2024", ocorrencia_raw="FERIAS", ocorrencia_tipo="ferias")]
    result = ExtractionResult(rows=rows, provider="gemini", pdf_type="native")
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
    ws = wb["Registros de Ponto"]
    fill = ws.cell(row=2, column=1).fill
    assert fill.fgColor.rgb.upper().endswith(OCORRENCIA_COLORS["ferias"].upper())


def test_resumo_has_provider():
    result = _make_result()
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
    ws = wb["Resumo"]
    values = [str(ws.cell(row=r, column=2).value or "") for r in range(1, 15)]
    assert "gemini" in values


def test_empty_result_does_not_crash():
    result = ExtractionResult(rows=[], provider="pdfplumber", pdf_type="native")
    output = build_excel(result)
    assert isinstance(output, bytes)


def test_all_occurrence_types_have_labels():
    from services.excel_builder import TIPO_LABELS
    for tipo in OCORRENCIA_COLORS:
        assert tipo in TIPO_LABELS, f"Missing label for {tipo}"


def test_data_row_count():
    rows = [TimesheetRow(data=f"0{i}/03/2024") for i in range(1, 6)]
    result = ExtractionResult(rows=rows, provider="pdfplumber", pdf_type="native")
    wb = openpyxl.load_workbook(io.BytesIO(build_excel(result)))
    ws = wb["Registros de Ponto"]
    # Row 1 is header, rows 2-6 are data
    assert ws.max_row == 6
