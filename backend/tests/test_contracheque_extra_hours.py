from __future__ import annotations

import asyncio
import io
import json
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl

from services.contracheque_extra_hours_excel_builder import (
    build_contracheque_extra_hours_excel,
)
from services.contracheque_extra_hours_service import (
    aggregate_extra_hours,
    is_extra_hour_description,
    stream_contracheque_extra_hours_extraction,
)

MINIMAL_PDF = (
    b"%PDF-1.4\n1 0 obj<</Type/Catalog>>endobj\n"
    b"xref\n0 2\ntrailer<</Size 2/Root 1 0 R>>\nstartxref\n9\n%%EOF"
)


def test_extra_hour_description_filter():
    assert is_extra_hour_description("Hora Extra Interjornada")
    assert is_extra_hour_description("HE Turno 100%")
    assert is_extra_hour_description("RSR-HE Troca de Turno")
    assert not is_extra_hour_description("Salario Basico")


def test_aggregate_extra_hours_sums_duplicates_and_preserves_new_columns():
    pages = [
        {
            "competencia": "04/2021",
            "itens": [
                {"descricao": "Hora Extra Interjornada", "valor": 498.79},
                {"descricao": "Salario Basico", "valor": 4292.77},
            ],
        },
        {
            "competencia": "05/2021",
            "itens": [
                {"descricao": "Hora Extra Interjornada", "valor": 100.0},
                {"descricao": "HE Turno 100%", "valor": 935.22},
                {"descricao": "HE Turno 100%", "valor": 10.0},
            ],
        },
    ]

    data, columns = aggregate_extra_hours(pages)

    assert columns == ["Hora Extra Interjornada", "HE Turno 100%"]
    assert data["04/2021"] == {"Hora Extra Interjornada": 498.79}
    assert data["05/2021"]["Hora Extra Interjornada"] == 100.0
    assert data["05/2021"]["HE Turno 100%"] == 945.22


def test_extra_hours_excel_has_dynamic_columns_blanks_and_total_formula():
    data = {
        "04/2021": {"Hora Extra Interjornada": 498.79},
        "05/2021": {"Hora Extra Interjornada": 100.0, "HE Turno 100%": 935.22},
    }
    columns = ["Hora Extra Interjornada", "HE Turno 100%"]

    wb = openpyxl.load_workbook(
        io.BytesIO(build_contracheque_extra_hours_excel(data, columns)),
        data_only=False,
    )
    ws = wb["Horas Extras"]

    assert [ws.cell(row=1, column=i).value for i in range(1, 5)] == [
        "Data",
        "Hora Extra Interjornada",
        "HE Turno 100%",
        "Total",
    ]
    assert ws.cell(row=2, column=1).value == "04/2021"
    assert ws.cell(row=2, column=2).value == 498.79
    assert ws.cell(row=2, column=3).value is None
    assert ws.cell(row=2, column=4).value == "=SUM(B2:C2)"
    assert ws.cell(row=3, column=1).value == "05/2021"
    assert ws.cell(row=3, column=4).value == "=SUM(B3:C3)"


def test_stream_skips_failed_pages_when_gemini_is_not_needed():
    page = {
        "competencia": "04/2021",
        "itens": [{"descricao": "Hora Extra Interjornada", "valor": 498.79}],
    }

    async def consume() -> dict:
        done = None
        async for event in stream_contracheque_extra_hours_extraction(MINIMAL_PDF, "test"):
            if event.startswith("data: "):
                payload = json.loads(event[6:])
                if payload.get("type") == "done":
                    done = payload
        assert done is not None
        return done

    with (
        patch(
            "services.contracheque_extra_hours_service._extract_all_pdfplumber",
            return_value=([page], [1]),
        ),
        patch(
            "services.contracheque_extra_hours_service._failed_pages_that_need_gemini",
            return_value=[],
        ),
        patch(
            "services.contracheque_extra_hours_service._process_chunk_gemini",
            AsyncMock(),
        ) as gemini_mock,
        patch(
            "services.contracheque_extra_hours_service.pypdf.PdfReader",
            return_value=MagicMock(pages=[object()]),
        ),
    ):
        result = asyncio.run(consume())

    assert result["provider"] == "pdfplumber"
    gemini_mock.assert_not_called()
